#!/usr/bin/env python3
"""
01_script_prepare_tabletask_metadata_qc_EG.py
By: Elnaz Ghasemi
Date: June 2026

Purpose
-------
Prepare the TableTask dataset metadata and QC files before running the
adapted Lena Movesense ECG and synchrony pipeline.

This script does NOT clean ECG, extract heart rate, calculate correlation,
calculate PLI, or run NSTE. It only creates the metadata and QC table that
later scripts will use.

Expected folder layout
----------------------
Run this script from the top-level TableTask folder:

TableTask/
  01_script_prepare_tabletask_metadata_qc_EG.py
  Table Task 2026.xlsx
  Table MoveSense Sensor/
  T24Z8B~S/
  table_timestamps.csv
  aggregated_data.csv

Main outputs
------------
qc_outputs/TableTask_Master_QC.xlsx
qc_outputs/veronica_analysis_units.csv
qc_outputs/veronica_analysis_units.xlsx
qc_outputs/metadata_qc_summary.txt

The Excel workbook contains several sheets. The most important sheet is
analysis_units. The standalone veronica_analysis_units.csv is the same table
saved as CSV so later Python scripts can read it easily.

Important design choices
------------------------
1. Candidate window 1 is practice. Cirkeline confirmed this.
2. Candidate windows 2 to 9 are task trials 1 to 8.
3. Trial timing comes from table_timestamps.csv rows where datatype == accel.
4. Trial condition labels come from Table Task 2026.xlsx, Trial Order.
5. Ball drop and other comments are flagged, but not excluded automatically.
6. Pilot sessions are flagged and excluded from main analysis by default.
7. Empty Movesense folders stay in QC summaries but are not used for analysis.
8. The script handles Excel-corrupted Participant N cells such as
   2204-01-01 00:00:00, which means participant 1, sensor suffix 2204.
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime, date, time, timezone, timedelta
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

from openpyxl.utils import get_column_letter
import numpy as np
import pandas as pd


# =============================================================================
# User-adjustable constants
# =============================================================================

# Local timezone for the study.
LOCAL_TZ = ZoneInfo("America/Vancouver")

# Movesense session folders are named like 2026_03_03-19_31_30.
SESSION_FOLDER_PATTERN = re.compile(r"^\d{4}_\d{2}_\d{2}-\d{2}_\d{2}_\d{2}$")

# Cirkeline confirmed: first accel-derived window is practice.
PRACTICE_CANDIDATE_WINDOW = 1

# Each real session should have practice plus 8 task trials.
EXPECTED_WINDOWS_PER_SESSION = 9

# Use this gap to split table accel windows into recording-level blocks.
# The table-task trial windows within one session are close together. The gap
# between sessions is much larger.
ACCEL_GROUP_GAP_MINUTES = 20

# Match an accel-window block to a Movesense folder if the first table-motion
# window starts within this many minutes of the Movesense folder local start.
# Some sessions start the table motion a few minutes after the Movesense start.
# Some practice windows can start slightly before one ECG file starts.
MAX_ACCEL_TO_FOLDER_START_DIFF_MINUTES = 15

# Pilot sessions are useful for testing but excluded from main analysis by default.
EXCLUDE_PILOTS_FROM_MAIN = True

# Comment flags are kept for sensitivity checks, but not excluded by default.
EXCLUDE_COMMENT_FLAGS_FROM_MAIN = False

# Known metadata issue:
# The metadata row for participants 23/24 is dated 2026/03/17 in the Excel file,
# but it matches the Movesense folder on 2026/03/27 at about 11:34 local time.
# This override keeps the raw date and also records the correction note.
KNOWN_METADATA_DATE_OVERRIDES = {
    ("P23_P24", "2026/03/17"): "2026/03/27",
}

# Known metadata sensor suffix typo:
# Participant 33 is written as sensor suffix 1204, but the matching Movesense
# folder has a sensor ending in 2204.
KNOWN_SENSOR_SUFFIX_CORRECTIONS = {
    "1204": "2204",
}

# Known table-timestamp repair:
# In the raw table_timestamps.csv, timestamp_id 39 is positioned immediately
# before timestamp_id 40 in the second March 3 afternoon block. Because its
# short accel time is 46:18.6-46:54.8, the generic Qualtrics-anchor inference
# can incorrectly place it at 14:46 and attach it to 2026_03_03-22_32_43.
# Manual inspection shows it belongs to the next session, 2026_03_03-23_45_18,
# at 15:46. This repair keeps March 3 sessions as 30-38 and 39-47.
KNOWN_ACCEL_TIME_REPAIRS = {
    39: {
        "date": "2026-03-03",
        "hour": 15,
        "note": "Manual repair: timestamp_id 39 belongs to the 2026_03_03-23_45_18 session at 15:46, not to the previous 14:32 session.",
    }
}

# Columns that should always be treated/displayed as text in Excel, because
# Excel otherwise displays long sensor IDs as scientific notation.
TEXT_DISPLAY_COLUMNS = {
    "sensor_id", "sensor_ids", "folder_sensor_ids",
    "sensor_A", "sensor_B",
    "sensor_1_suffix", "sensor_2_suffix",
    "sensor_A_suffix", "sensor_B_suffix",
    "sensor_suffix", "sensor_suffixes", "sensor_suffix_set",
    "sensor_1_suffix_original", "sensor_2_suffix_original",
}


# =============================================================================
# Basic utility functions
# =============================================================================

def parse_args() -> argparse.Namespace:
    """Read optional command-line arguments."""
    parser = argparse.ArgumentParser(
        description="Prepare TableTask metadata and QC files."
    )
    parser.add_argument(
        "--root",
        type=str,
        default=".",
        help="Top-level TableTask folder. Default: current working directory.",
    )
    return parser.parse_args()


def clean_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """Strip whitespace from column names."""
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def is_mac_metadata(path: Path) -> bool:
    """Return True for Mac metadata files/folders that should be ignored."""
    name = path.name
    return name == "__MACOSX" or name == ".DS_Store" or name.startswith("._")


def find_required_path(root: Path, candidates: list[str], label: str) -> Path:
    """Find a required input file/folder from a list of possible names."""
    for name in candidates:
        path = root / name
        if path.exists():
            return path
    tried = "\n  ".join(candidates)
    raise FileNotFoundError(f"Could not find {label}. Tried:\n  {tried}")


def find_metadata_path(root: Path) -> Path:
    """Find the Table Task metadata sheet."""
    candidates = [
        "Table Task 2026.xlsx",
        "Table Task 2026.xls",
        "Table Task 2026 - Sheet1.csv",
        "Table Task 2026.csv",
    ]
    for name in candidates:
        path = root / name
        if path.exists():
            return path

    # Fallback: use any spreadsheet/csv beginning with Table Task 2026.
    for path in sorted(root.glob("Table Task 2026*")):
        if path.suffix.lower() in {".xlsx", ".xls", ".csv"}:
            return path

    raise FileNotFoundError("Could not find Table Task 2026 metadata file.")


def session_folder_to_utc(session_name: str) -> datetime:
    """Parse a Movesense folder name as UTC datetime."""
    return datetime.strptime(session_name, "%Y_%m_%d-%H_%M_%S").replace(tzinfo=timezone.utc)


def session_folder_to_local(session_name: str) -> datetime:
    """Convert a Movesense folder name to Vancouver local datetime."""
    return session_folder_to_utc(session_name).astimezone(LOCAL_TZ)


def to_date_string(value: Any) -> str:
    """Convert a date-like value to YYYY/MM/DD text."""
    if pd.isna(value):
        return ""
    if isinstance(value, str) and value.strip().lower() == "pilot":
        return "pilot"
    try:
        return pd.to_datetime(value).strftime("%Y/%m/%d")
    except Exception:
        return str(value).strip()


def to_time_string(value: Any) -> str:
    """Convert a time-like value to HH:MM text."""
    if pd.isna(value):
        return ""
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    try:
        return pd.to_datetime(value).strftime("%H:%M")
    except Exception:
        return str(value).strip()


def contains_any(text: Any, keywords: list[str]) -> bool:
    """Return True if text contains any keyword, case-insensitive."""
    text_low = "" if pd.isna(text) else str(text).lower()
    return any(k.lower() in text_low for k in keywords)


def combine_text(*items: Any) -> str:
    """Combine non-empty text items with a separator."""
    parts = []
    for item in items:
        if pd.notna(item) and str(item).strip():
            parts.append(str(item).strip())
    return " | ".join(parts)


# =============================================================================
# Movesense raw data inventory and ECG coverage
# =============================================================================

def find_one_sensor_file(sensor_dir: Path, kind: str, session_name: str) -> Optional[Path]:
    """Find one ECG, Acc, or EulerAngles file inside a sensor folder.

    The full dataset usually has names like:
      ECG-2026_03_03-19_31_30.csv

    Earlier sample data sometimes had names like:
      251430002238_ECG-2026_03_03-19_31_30.csv

    The flexible pattern below handles both formats.
    """
    if kind == "ECG":
        patterns = [f"*ECG-{session_name}.csv", "*ECG-*.csv"]
    elif kind == "Acc":
        patterns = [f"*Acc-{session_name}.csv", "*Acc-*.csv"]
    elif kind == "EulerAngles":
        patterns = [f"*EulerAngles-{session_name}.csv", "*EulerAngles-*.csv"]
    else:
        raise ValueError(f"Unknown file kind: {kind}")

    matches: list[Path] = []
    for pattern in patterns:
        matches.extend([p for p in sensor_dir.glob(pattern) if not is_mac_metadata(p)])
    matches = sorted(set(matches))
    return matches[0] if matches else None


def find_annotation_file(session_dir: Path, session_name: str) -> Optional[Path]:
    """Find the annotation file for a Movesense session."""
    patterns = [f"Annotations-{session_name}.csv", "Annotations-*.csv"]
    matches: list[Path] = []
    for pattern in patterns:
        matches.extend([p for p in session_dir.glob(pattern) if not is_mac_metadata(p)])
    matches = sorted(set(matches))
    return matches[0] if matches else None


def estimate_csv_data_rows(path: Optional[Path]) -> Optional[int]:
    """Estimate data rows in a Movesense CSV.

    Movesense files have a title line, then a header line. This function is used
    only for inventory, not for numerical analysis.
    """
    if path is None or not path.exists():
        return None
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            n_lines = sum(1 for _ in f)
        return max(n_lines - 2, 0)
    except Exception:
        return None


def read_movesense_timestamps(path: Path) -> pd.Series:
    """Read the Timestamp column from a Movesense CSV.

    The real CSV header is on the second line, so skiprows=1 is required.
    """
    df = pd.read_csv(path, skiprows=1, usecols=["Timestamp"])
    return pd.to_numeric(df["Timestamp"], errors="coerce").dropna()


def get_ecg_coverage(ecg_file: Optional[Path]) -> dict[str, Any]:
    """Return ECG timing and sampling information for one sensor file."""
    if ecg_file is None or not ecg_file.exists():
        return {
            "ecg_present": False,
            "n_ecg_samples": np.nan,
            "first_ecg_unix": np.nan,
            "last_ecg_unix": np.nan,
            "first_ecg_local": "",
            "last_ecg_local": "",
            "ecg_duration_sec": np.nan,
            "median_ecg_dt_sec": np.nan,
            "approx_ecg_fs_hz": np.nan,
            "ecg_timestamps_monotonic": False,
            "ecg_backward_jumps": np.nan,
            "ecg_read_error": "",
        }

    try:
        ts = read_movesense_timestamps(ecg_file)
    except Exception as exc:
        return {
            "ecg_present": True,
            "n_ecg_samples": np.nan,
            "first_ecg_unix": np.nan,
            "last_ecg_unix": np.nan,
            "first_ecg_local": "",
            "last_ecg_local": "",
            "ecg_duration_sec": np.nan,
            "median_ecg_dt_sec": np.nan,
            "approx_ecg_fs_hz": np.nan,
            "ecg_timestamps_monotonic": False,
            "ecg_backward_jumps": np.nan,
            "ecg_read_error": str(exc),
        }

    if ts.empty:
        return {
            "ecg_present": True,
            "n_ecg_samples": 0,
            "first_ecg_unix": np.nan,
            "last_ecg_unix": np.nan,
            "first_ecg_local": "",
            "last_ecg_local": "",
            "ecg_duration_sec": np.nan,
            "median_ecg_dt_sec": np.nan,
            "approx_ecg_fs_hz": np.nan,
            "ecg_timestamps_monotonic": False,
            "ecg_backward_jumps": np.nan,
            "ecg_read_error": "Timestamp column was empty.",
        }

    # Use min/max for coverage because some files can have non-monotonic rows.
    first_ts = float(ts.min())
    last_ts = float(ts.max())
    duration = last_ts - first_ts

    # Use median positive timestamp difference for sampling-rate estimate.
    # This is more robust than n_samples / duration if there are timestamp jumps.
    diffs = ts.diff().dropna()
    backward_jumps = int((diffs < 0).sum())
    positive_diffs = diffs[diffs > 0]
    median_dt = float(positive_diffs.median()) if not positive_diffs.empty else np.nan
    approx_fs = float(1.0 / median_dt) if pd.notna(median_dt) and median_dt > 0 else np.nan

    first_local = pd.to_datetime(first_ts, unit="s", utc=True).tz_convert(str(LOCAL_TZ))
    last_local = pd.to_datetime(last_ts, unit="s", utc=True).tz_convert(str(LOCAL_TZ))

    return {
        "ecg_present": True,
        "n_ecg_samples": int(len(ts)),
        "first_ecg_unix": first_ts,
        "last_ecg_unix": last_ts,
        "first_ecg_local": str(first_local),
        "last_ecg_local": str(last_local),
        "ecg_duration_sec": duration,
        "median_ecg_dt_sec": median_dt,
        "approx_ecg_fs_hz": approx_fs,
        "ecg_timestamps_monotonic": bool(backward_jumps == 0),
        "ecg_backward_jumps": backward_jumps,
        "ecg_read_error": "",
    }


def build_movesense_inventory(movesense_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Scan the Table MoveSense Sensor folder and build raw data QC tables."""
    session_dirs = sorted(
        [
            p for p in movesense_dir.iterdir()
            if p.is_dir()
            and SESSION_FOLDER_PATTERN.match(p.name)
            and not is_mac_metadata(p)
        ],
        key=lambda p: p.name,
    )

    inventory_rows: list[dict[str, Any]] = []
    ecg_rows: list[dict[str, Any]] = []
    session_rows: list[dict[str, Any]] = []

    for session_dir in session_dirs:
        session_name = session_dir.name
        session_utc = session_folder_to_utc(session_name)
        session_local = session_utc.astimezone(LOCAL_TZ)
        annotation_file = find_annotation_file(session_dir, session_name)

        sensor_dirs = sorted(
            [p for p in session_dir.iterdir() if p.is_dir() and p.name.isdigit()],
            key=lambda p: p.name,
        )

        # Empty session folder. Keep it in session_summary, but it will not be
        # used for ECG analysis.
        if not sensor_dirs:
            session_rows.append(
                {
                    "recording_folder": session_name,
                    "recording_start_utc_from_folder": str(session_utc),
                    "recording_start_local_from_folder": str(session_local),
                    "n_sensor_folders": 0,
                    "sensor_ids": "",
                    "sensor_suffixes": "",
                    "annotation_file_present": annotation_file is not None,
                    "all_ecg_present": False,
                    "all_acc_present": False,
                    "all_euler_present": False,
                    "session_status": "empty_folder",
                    "session_complete_for_movesense": False,
                    "possible_ecg_dropout_by_duration": False,
                    "min_ecg_duration_sec": np.nan,
                    "max_ecg_duration_sec": np.nan,
                    "ecg_duration_range_sec": np.nan,
                    "session_notes": "No numeric sensor folders found.",
                }
            )
            continue

        sensor_ids: list[str] = []
        ecg_durations: list[float] = []
        session_has_any_sensor_file = False
        all_ecg_present = True
        all_acc_present = True
        all_euler_present = True

        for sensor_dir in sensor_dirs:
            sensor_id = sensor_dir.name
            sensor_ids.append(sensor_id)

            ecg_file = find_one_sensor_file(sensor_dir, "ECG", session_name)
            acc_file = find_one_sensor_file(sensor_dir, "Acc", session_name)
            euler_file = find_one_sensor_file(sensor_dir, "EulerAngles", session_name)

            # Some empty sessions contain sensor folders but no real sensor files.
            # Treat those as empty folders, not as ordinary missing-ECG sessions.
            if ecg_file is not None or acc_file is not None or euler_file is not None:
                session_has_any_sensor_file = True

            all_ecg_present = all_ecg_present and ecg_file is not None
            all_acc_present = all_acc_present and acc_file is not None
            all_euler_present = all_euler_present and euler_file is not None

            inventory_rows.append(
                {
                    "recording_folder": session_name,
                    "recording_start_utc_from_folder": str(session_utc),
                    "recording_start_local_from_folder": str(session_local),
                    "sensor_id": sensor_id,
                    "sensor_suffix": sensor_id[-4:],
                    "ecg_file_present": ecg_file is not None,
                    "acc_file_present": acc_file is not None,
                    "euler_file_present": euler_file is not None,
                    "annotation_file_present": annotation_file is not None,
                    "ecg_file": ecg_file.name if ecg_file else "",
                    "acc_file": acc_file.name if acc_file else "",
                    "euler_file": euler_file.name if euler_file else "",
                    "annotation_file": annotation_file.name if annotation_file else "",
                    "n_ecg_rows_estimate": estimate_csv_data_rows(ecg_file),
                    "n_acc_rows_estimate": estimate_csv_data_rows(acc_file),
                    "n_euler_rows_estimate": estimate_csv_data_rows(euler_file),
                }
            )

            ecg_info = get_ecg_coverage(ecg_file)
            ecg_rows.append(
                {
                    "recording_folder": session_name,
                    "sensor_id": sensor_id,
                    "sensor_suffix": sensor_id[-4:],
                    **ecg_info,
                }
            )

            if pd.notna(ecg_info["ecg_duration_sec"]):
                ecg_durations.append(float(ecg_info["ecg_duration_sec"]))

        min_duration = min(ecg_durations) if ecg_durations else np.nan
        max_duration = max(ecg_durations) if ecg_durations else np.nan
        duration_range = max_duration - min_duration if ecg_durations else np.nan

        possible_dropout = bool(pd.notna(duration_range) and duration_range > 300)

        if not session_has_any_sensor_file:
            status = "empty_folder"
            notes = "Sensor folders exist, but no ECG/Acc/Euler files were found."
        elif len(sensor_dirs) != 2:
            status = "check_sensor_count"
            notes = f"Expected 2 sensor folders, found {len(sensor_dirs)}."
        elif not all_ecg_present:
            status = "missing_ecg"
            notes = "At least one sensor is missing ECG."
        elif not (all_acc_present and all_euler_present):
            status = "missing_non_ecg_file"
            notes = "At least one sensor is missing Acc or EulerAngles."
        elif annotation_file is None:
            status = "missing_annotation"
            notes = "Annotation file is missing."
        else:
            status = "complete"
            notes = "Complete raw Movesense file set found."

        if possible_dropout:
            notes = notes + " ECG duration range suggests possible dropout."

        session_rows.append(
            {
                "recording_folder": session_name,
                "recording_start_utc_from_folder": str(session_utc),
                "recording_start_local_from_folder": str(session_local),
                "n_sensor_folders": len(sensor_dirs),
                "sensor_ids": ";".join(sensor_ids),
                "sensor_suffixes": ";".join([s[-4:] for s in sensor_ids]),
                "annotation_file_present": annotation_file is not None,
                "all_ecg_present": all_ecg_present,
                "all_acc_present": all_acc_present,
                "all_euler_present": all_euler_present,
                "session_status": status,
                "session_complete_for_movesense": status == "complete",
                "possible_ecg_dropout_by_duration": possible_dropout,
                "min_ecg_duration_sec": min_duration,
                "max_ecg_duration_sec": max_duration,
                "ecg_duration_range_sec": duration_range,
                "session_notes": notes,
            }
        )

    return pd.DataFrame(inventory_rows), pd.DataFrame(ecg_rows), pd.DataFrame(session_rows)


# =============================================================================
# table_timestamps.csv processing
# =============================================================================

def parse_full_local_datetime(value: Any) -> Optional[datetime]:
    """Parse a full Qualtrics timestamp as Vancouver local datetime."""
    if pd.isna(value):
        return None
    try:
        dt = pd.to_datetime(value)
        if pd.isna(dt):
            return None
        # Qualtrics rows in this file are local clock times with no reliable tz.
        return dt.to_pydatetime().replace(tzinfo=LOCAL_TZ)
    except Exception:
        return None


def parse_mmss_near_reference(reference: datetime, value: Any) -> Optional[datetime]:
    """Convert mm:ss.s to a full local datetime near a reference datetime."""
    try:
        text = str(value).strip()
        minute_text, second_text = text.split(":")
        minute = int(minute_text)
        second_float = float(second_text)
        second = int(second_float)
        microsecond = int(round((second_float - second) * 1_000_000))

        base = reference.replace(minute=minute, second=second, microsecond=microsecond)

        # Try previous, same, and next hour, then choose the closest to reference.
        candidates = [base + timedelta(hours=h) for h in [-1, 0, 1]]
        return min(candidates, key=lambda x: abs((x - reference).total_seconds()))
    except Exception:
        return None


def parse_mmss_with_fixed_date_hour(date_text: str, hour: int, value: Any) -> datetime:
    """Convert an mm:ss.s accel value using an explicit date and hour.

    This is used only for documented manual repairs where the generic nearest
    Qualtrics-anchor method is known to choose the wrong hour.
    """
    text = str(value).strip()
    minute_text, second_text = text.split(":")
    minute = int(minute_text)
    second_float = float(second_text)
    second = int(second_float)
    microsecond = int(round((second_float - second) * 1_000_000))
    fixed_date = pd.to_datetime(date_text).date()
    return datetime(
        fixed_date.year, fixed_date.month, fixed_date.day,
        hour, minute, second, microsecond, tzinfo=LOCAL_TZ
    )


def infer_accel_windows(table_timestamps_path: Path) -> pd.DataFrame:
    """Infer full datetimes for all accel rows in table_timestamps.csv."""
    raw = pd.read_csv(table_timestamps_path)
    raw = clean_column_names(raw)

    # The first column is the row ID in the original timestamp file.
    raw = raw.rename(columns={raw.columns[0]: "timestamp_id"})
    raw["datatype"] = raw["datatype"].astype(str).str.strip().str.lower()

    # Store full Qualtrics timestamps as anchors. We do not use Qualtrics for HR
    # segmentation, but it helps infer the date/hour for short accel mm:ss.s times.
    qual_rows: list[dict[str, Any]] = []
    for raw_index, row in raw.iterrows():
        if row["datatype"] != "qualtrics":
            continue
        q_start = parse_full_local_datetime(row["trial_start"])
        q_end = parse_full_local_datetime(row["trial_end"])
        if q_start is not None:
            qual_rows.append(
                {
                    "raw_index": raw_index,
                    "timestamp_id": row["timestamp_id"],
                    "q_start": q_start,
                    "q_end": q_end,
                }
            )

    accel_rows: list[dict[str, Any]] = []
    for raw_index, row in raw.iterrows():
        if row["datatype"] != "accel":
            continue

        # Use nearby Qualtrics rows first. If none are nearby, fall back to the
        # nearest Qualtrics row in file order.
        anchors = [q for q in qual_rows if abs(q["raw_index"] - raw_index) <= 12]
        if not anchors and qual_rows:
            anchors = [min(qual_rows, key=lambda q: abs(q["raw_index"] - raw_index))]

        best: Optional[dict[str, Any]] = None

        for anchor in anchors:
            start_local = parse_mmss_near_reference(anchor["q_start"], row["trial_start"])
            end_local = parse_mmss_near_reference(anchor["q_start"], row["trial_end"])

            if start_local is None or end_local is None:
                continue

            if end_local < start_local:
                end_local = end_local + timedelta(hours=1)

            # Score the inference. Lower score is better.
            # The inferred interval should be close to nearby Qualtrics time.
            time_score = min(
                abs((anchor["q_start"] - start_local).total_seconds()),
                abs((anchor["q_start"] - end_local).total_seconds()),
            )
            row_penalty = abs(anchor["raw_index"] - raw_index) * 5
            score = time_score + row_penalty

            candidate = {
                "start_local": start_local,
                "end_local": end_local,
                "anchor_qualtrics_row_index": anchor["raw_index"],
                "anchor_qualtrics_start": anchor["q_start"],
                "accel_time_inference_score": score,
            }

            if best is None or candidate["accel_time_inference_score"] < best["accel_time_inference_score"]:
                best = candidate

        if best is None:
            start_local = pd.NaT
            end_local = pd.NaT
            start_unix = np.nan
            end_unix = np.nan
            duration = np.nan
            note = "Could not infer full datetime."
            anchor_index = np.nan
            anchor_start = ""
            score = np.nan
        else:
            start_local = best["start_local"]
            end_local = best["end_local"]
            start_unix = start_local.astimezone(timezone.utc).timestamp()
            end_unix = end_local.astimezone(timezone.utc).timestamp()
            duration = end_unix - start_unix
            note = "Inferred from nearby Qualtrics timestamp."
            anchor_index = best["anchor_qualtrics_row_index"]
            anchor_start = str(best["anchor_qualtrics_start"])
            score = best["accel_time_inference_score"]

        # Apply documented manual accel-time repair after the generic inference.
        # This preserves the original row and raw mm:ss values, but corrects the
        # full local/Unix timestamp used for grouping and session matching.
        try:
            timestamp_id_int = int(row["timestamp_id"])
        except Exception:
            timestamp_id_int = None

        repair_applied = False
        repair_note = ""
        if timestamp_id_int in KNOWN_ACCEL_TIME_REPAIRS:
            repair = KNOWN_ACCEL_TIME_REPAIRS[timestamp_id_int]
            start_local = parse_mmss_with_fixed_date_hour(repair["date"], repair["hour"], row["trial_start"])
            end_local = parse_mmss_with_fixed_date_hour(repair["date"], repair["hour"], row["trial_end"])
            if end_local < start_local:
                end_local = end_local + timedelta(hours=1)
            start_unix = start_local.astimezone(timezone.utc).timestamp()
            end_unix = end_local.astimezone(timezone.utc).timestamp()
            duration = end_unix - start_unix
            note = note + " | Manual documented repair applied."
            repair_applied = True
            repair_note = repair["note"]

        accel_rows.append(
            {
                "table_timestamps_row_index": raw_index,
                "timestamp_id": row["timestamp_id"],
                "accel_start_raw": row["trial_start"],
                "accel_end_raw": row["trial_end"],
                "accel_start_local": str(start_local) if not pd.isna(start_local) else "",
                "accel_end_local": str(end_local) if not pd.isna(end_local) else "",
                "accel_start_unix": start_unix,
                "accel_end_unix": end_unix,
                "window_duration_sec": duration,
                "anchor_qualtrics_row_index": anchor_index,
                "anchor_qualtrics_start": anchor_start,
                "accel_time_inference_score": score,
                "accel_time_inference_note": note,
                "accel_time_repair_applied": repair_applied,
                "accel_time_repair_note": repair_note,
            }
        )

    out = pd.DataFrame(accel_rows)
    out = out.sort_values("accel_start_unix", na_position="last").reset_index(drop=True)

    # Split accel rows into session-level blocks by large time gaps.
    out["previous_gap_min"] = out["accel_start_unix"].diff() / 60
    out["accel_group_id"] = (out["previous_gap_min"].fillna(9999) > ACCEL_GROUP_GAP_MINUTES).cumsum()

    # Candidate window numbering is within each accel group.
    out["candidate_window"] = out.groupby("accel_group_id").cumcount() + 1
    out["trial"] = out["candidate_window"] - 1
    out["is_practice"] = out["candidate_window"] == PRACTICE_CANDIDATE_WINDOW

    return out


def summarize_accel_groups(accel_windows: pd.DataFrame) -> pd.DataFrame:
    """Create one row per inferred table-accelerometer recording block."""
    rows: list[dict[str, Any]] = []
    for group_id, group in accel_windows.groupby("accel_group_id", sort=True):
        group = group.sort_values("accel_start_unix")
        start_unix = float(group["accel_start_unix"].iloc[0])
        end_unix = float(group["accel_end_unix"].iloc[-1])
        start_local = pd.to_datetime(start_unix, unit="s", utc=True).tz_convert(str(LOCAL_TZ))
        end_local = pd.to_datetime(end_unix, unit="s", utc=True).tz_convert(str(LOCAL_TZ))
        rows.append(
            {
                "accel_group_id": int(group_id),
                "n_windows_in_group": int(len(group)),
                "group_start_unix": start_unix,
                "group_end_unix": end_unix,
                "group_start_local": str(start_local),
                "group_end_local": str(end_local),
                "group_date": start_local.strftime("%Y/%m/%d"),
                "group_start_time": start_local.strftime("%H:%M:%S"),
                "timestamp_ids": ";".join(map(str, group["timestamp_id"].tolist())),
                "group_note": (
                    "Expected 9 windows."
                    if len(group) == EXPECTED_WINDOWS_PER_SESSION
                    else f"Expected 9 windows, found {len(group)}."
                ),
            }
        )
    return pd.DataFrame(rows)


# =============================================================================
# Metadata parsing
# =============================================================================

def read_metadata_sheet(path: Path) -> pd.DataFrame:
    """Read Table Task metadata from XLSX or CSV."""
    if path.suffix.lower() in {".xlsx", ".xls"}:
        df = pd.read_excel(path)
    elif path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        raise ValueError(f"Unsupported metadata file type: {path.suffix}")
    return clean_column_names(df)


def parse_participant_cell(value: Any) -> dict[str, Any]:
    """Parse Participant N cell.

    Supported formats:
    1. Normal text:
       "17, 2277" means participant 17, sensor suffix 2277.

    2. Text with role notes:
       "6, 2208 (confederate)" means participant 6, sensor suffix 2208.

    3. Excel-corrupted early cells:
       datetime(2204, 1, 1) means participant 1, sensor suffix 2204.
       datetime(2208, 10, 1) means participant 10, sensor suffix 2208.
    """
    is_confederate = False

    # Excel date-corrupted format.
    if isinstance(value, (pd.Timestamp, datetime)):
        sensor_suffix = str(value.year)
        participant = int(value.month)
        return {
            "participant": participant,
            "sensor_suffix_original": sensor_suffix,
            "sensor_suffix": KNOWN_SENSOR_SUFFIX_CORRECTIONS.get(sensor_suffix, sensor_suffix),
            "participant_cell_format": "excel_corrupted_date",
            "is_confederate": False,
            "raw_participant_cell": str(value),
        }

    # Normal text format.
    text = "" if pd.isna(value) else str(value).strip()
    is_confederate = "confederate" in text.lower()
    match = re.search(r"(\d+)\s*,\s*(\d+)", text)

    if match:
        participant = int(match.group(1))
        suffix_original = str(match.group(2))
        suffix_corrected = KNOWN_SENSOR_SUFFIX_CORRECTIONS.get(suffix_original, suffix_original)
        return {
            "participant": participant,
            "sensor_suffix_original": suffix_original,
            "sensor_suffix": suffix_corrected,
            "participant_cell_format": "text_comma",
            "is_confederate": is_confederate,
            "raw_participant_cell": text,
        }

    # Unparseable cell.
    return {
        "participant": np.nan,
        "sensor_suffix_original": "",
        "sensor_suffix": "",
        "participant_cell_format": "unparsed",
        "is_confederate": is_confederate,
        "raw_participant_cell": text,
    }


def normalize_metadata_date(value: Any) -> str:
    """Normalize metadata date, preserving pilot marker."""
    if pd.isna(value):
        return ""
    if isinstance(value, str) and value.strip().lower() == "pilot":
        return "pilot"
    try:
        return pd.to_datetime(value).strftime("%Y/%m/%d")
    except Exception:
        return str(value).strip()


def metadata_time_candidates(date_text: str, time_text: str) -> list[datetime]:
    """Return plausible local datetimes for metadata date/time.

    Excel stores afternoon sessions as 02:30, 03:30, etc. Those can mean AM
    literally, but in this study they often mean PM. We return both the literal
    time and a PM-adjusted time when the hour is 1 to 6.
    """
    if not date_text or date_text == "pilot" or not time_text:
        return []

    try:
        date_part = pd.to_datetime(date_text).date()
    except Exception:
        return []

    try:
        parsed = pd.to_datetime(time_text).time()
    except Exception:
        return []

    candidates: list[datetime] = []
    literal = datetime.combine(date_part, parsed).replace(tzinfo=LOCAL_TZ)
    candidates.append(literal)

    # Add PM version for ambiguous early-hour metadata times.
    if 1 <= parsed.hour <= 6:
        pm = literal + timedelta(hours=12)
        candidates.append(pm)

    # Remove duplicates while preserving order.
    unique: list[datetime] = []
    seen = set()
    for dt in candidates:
        key = dt.isoformat()
        if key not in seen:
            unique.append(dt)
            seen.add(key)

    return unique


def apply_known_metadata_date_override(dyad_id: str, date_text: str) -> tuple[str, str]:
    """Apply known metadata date correction, returning corrected date and note."""
    key = (dyad_id, date_text)
    if key in KNOWN_METADATA_DATE_OVERRIDES:
        return KNOWN_METADATA_DATE_OVERRIDES[key], f"Known date override: {date_text} -> {KNOWN_METADATA_DATE_OVERRIDES[key]}"
    return date_text, ""


def build_metadata_dyads(metadata: pd.DataFrame, aggregated: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Convert participant-row metadata into one row per dyad/session.

    The Excel sheet uses two rows per dyad. The first row contains date, time,
    trial order, and participant 1. The next row contains participant 2 and
    sometimes pilot status or comments.
    """
    required_cols = ["Date", "Time", "Participant N", "A/B", "Trial Order", "Comments", "RA"]
    missing = [c for c in required_cols if c not in metadata.columns]
    if missing:
        raise ValueError(f"Metadata sheet is missing required columns: {missing}")

    raw_blocks: list[dict[str, Any]] = []
    dyad_rows: list[dict[str, Any]] = []

    i = 0
    while i < len(metadata):
        row1 = metadata.iloc[i]

        # Dyad starts when Trial Order is present.
        if pd.isna(row1.get("Trial Order")) or str(row1.get("Trial Order")).strip() == "":
            i += 1
            continue

        if i + 1 >= len(metadata):
            break

        row2 = metadata.iloc[i + 1]

        p1 = parse_participant_cell(row1.get("Participant N"))
        p2 = parse_participant_cell(row2.get("Participant N"))

        role1 = "" if pd.isna(row1.get("A/B")) else str(row1.get("A/B")).strip().upper()
        role2 = "" if pd.isna(row2.get("A/B")) else str(row2.get("A/B")).strip().upper()

        date_raw = normalize_metadata_date(row1.get("Date"))
        time_text = to_time_string(row1.get("Time"))

        # Construct dyad ID early so known overrides can use it.
        if pd.notna(p1["participant"]) and pd.notna(p2["participant"]):
            dyad_id = f"P{int(p1['participant']):02d}_P{int(p2['participant']):02d}"
        else:
            dyad_id = ""

        date_corrected, date_note = apply_known_metadata_date_override(dyad_id, date_raw)
        time_candidates = metadata_time_candidates(date_corrected, time_text)

        comments = combine_text(row1.get("Comments"), row2.get("Comments"))
        ra = row1.get("RA") if pd.notna(row1.get("RA")) else row2.get("RA")

        is_pilot = (
            str(row1.get("Date", "")).strip().lower() == "pilot"
            or str(row2.get("Date", "")).strip().lower() == "pilot"
            or contains_any(row1.get("Comments", ""), ["pilot"])
            or contains_any(row2.get("Comments", ""), ["pilot"])
        )

        participants = [
            {**p1, "role": role1},
            {**p2, "role": role2},
        ]

        A = next((p for p in participants if p["role"] == "A"), None)
        B = next((p for p in participants if p["role"] == "B"), None)

        participant_values = [p1["participant"], p2["participant"]]
        participant_values = [p for p in participant_values if pd.notna(p)]
        participant_mean = float(np.mean(participant_values)) if participant_values else np.nan

        pair1 = infer_pair1_from_aggregated(aggregated, participant_mean)

        raw_blocks.append(
            {
                "metadata_row_1_index": int(i),
                "metadata_row_2_index": int(i + 1),
                "date_raw": str(row1.get("Date")),
                "date_normalized": date_raw,
                "date_for_matching": date_corrected,
                "time_raw": str(row1.get("Time")),
                "time_normalized": time_text,
                "participant_1_raw": p1["raw_participant_cell"],
                "participant_2_raw": p2["raw_participant_cell"],
                "participant_1_format": p1["participant_cell_format"],
                "participant_2_format": p2["participant_cell_format"],
                "dyad_id": dyad_id,
                "sensor_suffix_set": ";".join(sorted([p1["sensor_suffix"], p2["sensor_suffix"]])),
                "trial_order": row1.get("Trial Order"),
                "is_pilot": is_pilot,
                "date_correction_note": date_note,
            }
        )

        dyad_rows.append(
            {
                "metadata_block_index": len(raw_blocks),
                "metadata_row_1_index": int(i),
                "metadata_row_2_index": int(i + 1),
                "session_date_raw": date_raw,
                "session_date": date_corrected,
                "session_time": time_text,
                "metadata_time_candidates_local": ";".join([str(x) for x in time_candidates]),
                "dyad_id": dyad_id,
                "participant_1": p1["participant"],
                "participant_2": p2["participant"],
                "participant_mean": participant_mean,
                "pair1": pair1,
                "sensor_1_suffix": p1["sensor_suffix"],
                "sensor_2_suffix": p2["sensor_suffix"],
                "sensor_1_suffix_original": p1["sensor_suffix_original"],
                "sensor_2_suffix_original": p2["sensor_suffix_original"],
                "sensor_suffix_set": ";".join(sorted([p1["sensor_suffix"], p2["sensor_suffix"]])),
                "participant_A": A["participant"] if A else np.nan,
                "participant_B": B["participant"] if B else np.nan,
                "sensor_A_suffix": A["sensor_suffix"] if A else "",
                "sensor_B_suffix": B["sensor_suffix"] if B else "",
                "participant_A_is_confederate": A["is_confederate"] if A else False,
                "participant_B_is_confederate": B["is_confederate"] if B else False,
                "trial_order": row1.get("Trial Order"),
                "is_pilot": is_pilot,
                "session_comments": comments,
                "RA": ra,
                "ball_drop_flag": contains_any(comments, ["ball fell", "ball dropped", "ball drop", "bell fell"]),
                "stopped_flag": contains_any(comments, ["stopped", "stoped", "almost stopped"]),
                "talking_laughing_flag": contains_any(comments, ["talking", "talked", "laugh", "chit-chat", "chatted"]),
                "wrong_way_flag": contains_any(comments, ["wrong way", "went over finish", "finish line"]),
                "sensor_issue_comment_flag": contains_any(comments, ["sensor", "movesense", "recording"]),
                "date_correction_note": date_note,
            }
        )

        i += 2

    return pd.DataFrame(dyad_rows), pd.DataFrame(raw_blocks)


def infer_pair1_from_aggregated(aggregated: pd.DataFrame, participant_mean: float) -> Any:
    """Infer pair1 from aggregated_data.csv using Participant_ID_mean."""
    if pd.isna(participant_mean) or "Participant_ID_mean" not in aggregated.columns:
        return np.nan

    temp = aggregated.copy()
    temp = temp[pd.notna(temp["Participant_ID_mean"])]

    # Trial 0 can be practice and can sometimes behave differently. Use task
    # trials to infer the pair index when possible.
    if "trial" in temp.columns:
        temp_task = temp[temp["trial"] > 0]
        if not temp_task.empty:
            temp = temp_task

    if temp.empty:
        return np.nan

    summary = (
        temp.groupby("pair1", as_index=False)["Participant_ID_mean"]
        .median()
        .rename(columns={"Participant_ID_mean": "participant_id_mean_median"})
    )
    summary["diff"] = (summary["participant_id_mean_median"] - participant_mean).abs()
    best = summary.sort_values("diff").iloc[0]

    # Tolerance allows small aggregation irregularities.
    if best["diff"] <= 0.6:
        return int(best["pair1"])
    return np.nan


def expand_trial_order(order: Any) -> dict[int, str]:
    """Expand a 4-block trial order into 8 trial-level labels."""
    if pd.isna(order):
        return {}

    blocks = [b.strip().upper() for b in str(order).split("-") if b.strip()]
    expanded: list[str] = []

    # Each block is done twice, giving 8 task trials.
    for block in blocks:
        expanded.extend([block, block])

    return {trial_number: condition for trial_number, condition in enumerate(expanded, start=1)}


# =============================================================================
# Matching metadata and accel groups to Movesense folders
# =============================================================================

def candidate_time_difference_seconds(folder_local: datetime, metadata_row: pd.Series) -> float:
    """Return smallest difference between folder time and metadata time candidates."""
    candidates_text = metadata_row.get("metadata_time_candidates_local", "")
    if not candidates_text:
        return float("inf")

    best = float("inf")
    for piece in str(candidates_text).split(";"):
        piece = piece.strip()
        if not piece:
            continue
        try:
            dt = pd.to_datetime(piece).to_pydatetime()
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=LOCAL_TZ)
            diff = abs((dt - folder_local).total_seconds())
            best = min(best, diff)
        except Exception:
            continue
    return best


def sensor_set_overlap(folder_suffixes: set[str], meta_suffixes: set[str]) -> int:
    """Count overlapping sensor suffixes between folder and metadata."""
    return len(folder_suffixes.intersection(meta_suffixes))


def resolve_sensor_roles(sensor_ids: list[str], metadata_row: pd.Series) -> tuple[str, str, str]:
    """Resolve full sensor IDs for A and B roles.

    Exact suffix matching is used first. If a metadata suffix is corrected or
    mistyped and only one sensor remains, assign the remaining sensor and record
    the note.
    """
    note_parts: list[str] = []
    sensor_A = ""
    sensor_B = ""

    suffix_A = str(metadata_row.get("sensor_A_suffix", ""))
    suffix_B = str(metadata_row.get("sensor_B_suffix", ""))

    for sensor in sensor_ids:
        if suffix_A and sensor.endswith(suffix_A):
            sensor_A = sensor
        if suffix_B and sensor.endswith(suffix_B):
            sensor_B = sensor

    # If one role was not matched but there are exactly two sensors, infer it as
    # the remaining sensor. This handles known suffix typo 1204 -> 2204.
    if len(sensor_ids) == 2:
        if sensor_A and not sensor_B:
            remaining = [s for s in sensor_ids if s != sensor_A]
            if len(remaining) == 1:
                sensor_B = remaining[0]
                note_parts.append("Inferred sensor_B as remaining folder sensor.")
        elif sensor_B and not sensor_A:
            remaining = [s for s in sensor_ids if s != sensor_B]
            if len(remaining) == 1:
                sensor_A = remaining[0]
                note_parts.append("Inferred sensor_A as remaining folder sensor.")

    if not sensor_A:
        note_parts.append("Could not resolve sensor_A.")
    if not sensor_B:
        note_parts.append("Could not resolve sensor_B.")

    return sensor_A, sensor_B, " ".join(note_parts)


def match_metadata_to_folders(
    metadata_dyads: pd.DataFrame,
    session_summary: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Match each Movesense session folder to the best metadata dyad row."""
    candidate_rows: list[dict[str, Any]] = []
    match_rows: list[dict[str, Any]] = []
    unmatched_rows: list[dict[str, Any]] = []

    used_metadata_blocks: set[int] = set()

    # Evaluate match candidates for all folders.
    for _, folder_row in session_summary.iterrows():
        folder = folder_row["recording_folder"]
        folder_local = session_folder_to_local(folder)
        folder_date = folder_local.strftime("%Y/%m/%d")
        sensor_ids = [s for s in str(folder_row.get("sensor_ids", "")).split(";") if s]
        folder_suffixes = {s[-4:] for s in sensor_ids}

        for _, meta in metadata_dyads.iterrows():
            meta_suffixes = {str(meta.get("sensor_1_suffix", "")), str(meta.get("sensor_2_suffix", ""))}
            meta_suffixes = {s for s in meta_suffixes if s}

            same_date = meta.get("session_date", "") == folder_date
            overlap = sensor_set_overlap(folder_suffixes, meta_suffixes)
            exact_sensor_set = folder_suffixes == meta_suffixes if folder_suffixes and meta_suffixes else False
            time_diff = candidate_time_difference_seconds(folder_local, meta)

            # Candidate logic:
            # - Prefer same date and exact sensor suffix set.
            # - Allow one-sensor overlap if time is very close, for known suffix typo.
            # - Allow date mismatch only if time-of-day and sensors strongly match,
            #   for known metadata date typo P23/P24.
            candidate_type = ""
            if same_date and exact_sensor_set:
                candidate_type = "same_date_exact_sensors"
                score = time_diff
            elif same_date and overlap >= 1 and time_diff <= 20 * 60:
                candidate_type = "same_date_partial_sensor_time_close"
                score = time_diff + 500
            elif (not same_date) and exact_sensor_set and time_diff <= 20 * 60:
                candidate_type = "date_mismatch_exact_sensors_time_close"
                score = time_diff + 2000
            else:
                continue

            candidate_rows.append(
                {
                    "recording_folder": folder,
                    "recording_start_local_from_folder": str(folder_local),
                    "folder_sensor_ids": ";".join(sensor_ids),
                    "folder_sensor_suffixes": ";".join(sorted(folder_suffixes)),
                    "folder_session_status": folder_row.get("session_status", ""),
                    "metadata_block_index": meta.get("metadata_block_index"),
                    "dyad_id": meta.get("dyad_id", ""),
                    "metadata_session_date": meta.get("session_date", ""),
                    "metadata_session_time": meta.get("session_time", ""),
                    "metadata_sensor_suffix_set": meta.get("sensor_suffix_set", ""),
                    "candidate_type": candidate_type,
                    "same_date": same_date,
                    "exact_sensor_set": exact_sensor_set,
                    "sensor_suffix_overlap_count": overlap,
                    "time_diff_sec": time_diff,
                    "match_score": score,
                }
            )

    candidates = pd.DataFrame(candidate_rows)

    # Choose one metadata row per folder, and avoid reusing the same metadata row
    # when possible.
    if not candidates.empty:
        candidates = candidates.sort_values(["recording_folder", "match_score"]).reset_index(drop=True)

    for _, folder_row in session_summary.sort_values("recording_folder").iterrows():
        folder = folder_row["recording_folder"]
        folder_candidates = candidates[candidates["recording_folder"] == folder].copy() if not candidates.empty else pd.DataFrame()

        if folder_candidates.empty:
            unmatched_rows.append(
                {
                    "recording_folder": folder,
                    "recording_start_local_from_folder": str(session_folder_to_local(folder)),
                    "sensor_ids": folder_row.get("sensor_ids", ""),
                    "session_status": folder_row.get("session_status", ""),
                    "matching_issue": "No candidate metadata row matched date/time/sensors.",
                }
            )
            continue

        # Prefer unused metadata blocks. If all are used, use the best one but
        # record the reuse. This should be rare.
        folder_candidates["metadata_already_used"] = folder_candidates["metadata_block_index"].isin(used_metadata_blocks)
        unused_candidates = folder_candidates[~folder_candidates["metadata_already_used"]]
        if not unused_candidates.empty:
            chosen = unused_candidates.sort_values("match_score").iloc[0]
            reuse_note = ""
        else:
            chosen = folder_candidates.sort_values("match_score").iloc[0]
            reuse_note = "Metadata block was already used by another folder."

        block_index = int(chosen["metadata_block_index"])
        used_metadata_blocks.add(block_index)

        meta = metadata_dyads[metadata_dyads["metadata_block_index"] == block_index].iloc[0]
        sensor_ids = [s for s in str(folder_row.get("sensor_ids", "")).split(";") if s]
        sensor_A, sensor_B, role_note = resolve_sensor_roles(sensor_ids, meta)

        match_rows.append(
            {
                **meta.to_dict(),
                "recording_folder": folder,
                "recording_start_local_from_folder": str(session_folder_to_local(folder)),
                "folder_sensor_ids": ";".join(sensor_ids),
                "folder_session_status": folder_row.get("session_status", ""),
                "folder_metadata_match_note": f"{chosen['candidate_type']}. {reuse_note}".strip(),
                "metadata_match_time_diff_sec": chosen["time_diff_sec"],
                "metadata_match_score": chosen["match_score"],
                "sensor_A": sensor_A,
                "sensor_B": sensor_B,
                "sensor_role_resolution_note": role_note,
            }
        )

    unmatched = pd.DataFrame(unmatched_rows)
    matches = pd.DataFrame(match_rows)

    return matches, unmatched, candidates


def match_accel_groups_to_folders(
    accel_groups: pd.DataFrame,
    session_summary: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Match each accel group to the closest Movesense folder by local time."""
    rows: list[dict[str, Any]] = []
    unmatched_groups: list[dict[str, Any]] = []

    # Keep all folders in the matching table, including empty folders. Empty
    # folders can match an accel group, but they will not be used for analysis.
    session_records = []
    for _, session in session_summary.iterrows():
        folder_local = session_folder_to_local(session["recording_folder"])
        session_records.append(
            {
                **session.to_dict(),
                "folder_local_dt_obj": folder_local,
            }
        )

    for _, group in accel_groups.iterrows():
        group_start = pd.to_datetime(group["group_start_unix"], unit="s", utc=True).tz_convert(str(LOCAL_TZ)).to_pydatetime()
        if group_start.tzinfo is None:
            group_start = group_start.replace(tzinfo=LOCAL_TZ)

        candidates: list[dict[str, Any]] = []
        for session in session_records:
            diff_min = abs((group_start - session["folder_local_dt_obj"]).total_seconds()) / 60
            if diff_min <= MAX_ACCEL_TO_FOLDER_START_DIFF_MINUTES:
                candidates.append(
                    {
                        **session,
                        "accel_group_id": group["accel_group_id"],
                        "group_start_local": group["group_start_local"],
                        "n_windows_in_group": group["n_windows_in_group"],
                        "accel_to_folder_start_diff_min": diff_min,
                    }
                )

        if not candidates:
            unmatched_groups.append(
                {
                    "accel_group_id": group["accel_group_id"],
                    "group_start_local": group["group_start_local"],
                    "n_windows_in_group": group["n_windows_in_group"],
                    "matching_issue": "No Movesense folder start time within threshold.",
                }
            )
            continue

        chosen = sorted(candidates, key=lambda x: x["accel_to_folder_start_diff_min"])[0]
        rows.append(chosen)

    return pd.DataFrame(rows), pd.DataFrame(unmatched_groups)


def build_candidate_window_coverage(
    accel_windows: pd.DataFrame,
    accel_group_folder_matches: pd.DataFrame,
    session_summary: pd.DataFrame,
    ecg_qc: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Create one row per matched folder-window and check ECG coverage."""
    rows: list[dict[str, Any]] = []
    unmatched_complete_sessions: list[dict[str, Any]] = []

    matched_group_to_folder = {
        int(row["accel_group_id"]): row.to_dict()
        for _, row in accel_group_folder_matches.iterrows()
    }

    ecg_by_folder = {folder: df.copy() for folder, df in ecg_qc.groupby("recording_folder")}

    for group_id, group in accel_windows.groupby("accel_group_id", sort=True):
        group_id = int(group_id)
        folder_match = matched_group_to_folder.get(group_id)
        if folder_match is None:
            continue

        folder = folder_match["recording_folder"]
        folder_status = folder_match.get("session_status", "")
        session_complete = bool(folder_match.get("session_complete_for_movesense", False))
        session_ecg = ecg_by_folder.get(folder, pd.DataFrame())

        # Keep candidate-window QC only for complete Movesense sessions. Empty
        # folders are represented in session_summary and accel_group_folder_matches.
        if not session_complete:
            continue

        group = group.sort_values("accel_start_unix")

        # Some inferred accel groups have 10 rows. The experimental structure is
        # practice plus 8 real trials, so only the first 9 windows are used here.
        # The full group remains visible in accel_groups and inferred_accel_windows.
        if len(group) > EXPECTED_WINDOWS_PER_SESSION:
            selected_group = group.head(EXPECTED_WINDOWS_PER_SESSION).copy()
            group_selection_note = (
                f"Group had {len(group)} windows; selected first {EXPECTED_WINDOWS_PER_SESSION}."
            )
        else:
            selected_group = group.copy()
            group_selection_note = f"Group had {len(group)} windows."

        for _, w in selected_group.iterrows():
            start_unix = float(w["accel_start_unix"])
            end_unix = float(w["accel_end_unix"])

            covered_count = 0
            coverage_values: dict[str, Any] = {}

            for _, sensor_row in session_ecg.iterrows():
                sensor_id = str(sensor_row["sensor_id"])
                first_ecg = sensor_row.get("first_ecg_unix", np.nan)
                last_ecg = sensor_row.get("last_ecg_unix", np.nan)
                covered = bool(
                    pd.notna(first_ecg)
                    and pd.notna(last_ecg)
                    and float(first_ecg) <= start_unix
                    and float(last_ecg) >= end_unix
                )
                coverage_values[f"sensor_{sensor_id}_covered"] = covered
                if covered:
                    covered_count += 1

            usable = bool(covered_count >= 2)

            rows.append(
                {
                    "recording_folder": folder,
                    "folder_session_status": folder_status,
                    "accel_group_id": group_id,
                    "candidate_window": int(w["candidate_window"]),
                    "trial": int(w["trial"]),
                    "is_practice": bool(w["is_practice"]),
                    "timestamp_id": w["timestamp_id"],
                    "table_timestamps_row_index": w["table_timestamps_row_index"],
                    "accel_start_raw": w["accel_start_raw"],
                    "accel_end_raw": w["accel_end_raw"],
                    "accel_start_local": w["accel_start_local"],
                    "accel_end_local": w["accel_end_local"],
                    "accel_start_unix": start_unix,
                    "accel_end_unix": end_unix,
                    "window_duration_sec": w["window_duration_sec"],
                    "covered_sensor_count": covered_count,
                    "usable_for_dyadic_ecg": usable,
                    "n_windows_in_accel_group": int(len(group)),
                    "n_windows_selected_for_analysis": int(len(selected_group)),
                    "window_group_selection_note": group_selection_note,
                    "accel_to_folder_start_diff_min": folder_match.get("accel_to_folder_start_diff_min", np.nan),
                    **coverage_values,
                }
            )

    candidate_window_coverage = pd.DataFrame(rows)

    # Complete sessions without windows are important QC. Example: one Movesense
    # session can exist without matching table accelerometer windows.
    folders_with_windows = set(candidate_window_coverage["recording_folder"]) if not candidate_window_coverage.empty else set()
    for _, session in session_summary.iterrows():
        if bool(session.get("session_complete_for_movesense", False)) and session["recording_folder"] not in folders_with_windows:
            unmatched_complete_sessions.append(
                {
                    "recording_folder": session["recording_folder"],
                    "recording_start_local_from_folder": session["recording_start_local_from_folder"],
                    "sensor_ids": session["sensor_ids"],
                    "session_status": session["session_status"],
                    "matching_issue": "Complete Movesense session has no matched accel-window group.",
                }
            )

    return candidate_window_coverage, pd.DataFrame(unmatched_complete_sessions)


# =============================================================================
# Analysis units
# =============================================================================

def build_analysis_units(
    candidate_windows: pd.DataFrame,
    folder_metadata: pd.DataFrame,
    aggregated: pd.DataFrame,
) -> pd.DataFrame:
    """Merge windows, metadata, and aggregated data into one master table."""
    rows: list[dict[str, Any]] = []
    folder_meta = {
        row["recording_folder"]: row.to_dict()
        for _, row in folder_metadata.iterrows()
    }

    for _, w in candidate_windows.iterrows():
        folder = w["recording_folder"]
        meta = folder_meta.get(folder)

        if meta is None:
            # Keep row but flag missing metadata.
            meta = {
                "recording_folder": folder,
                "dyad_id": "",
                "pair1": np.nan,
                "participant_A": np.nan,
                "participant_B": np.nan,
                "sensor_A": "",
                "sensor_B": "",
                "trial_order": "",
                "is_pilot": False,
                "session_comments": "",
                "RA": "",
                "ball_drop_flag": False,
                "stopped_flag": False,
                "talking_laughing_flag": False,
                "wrong_way_flag": False,
                "sensor_issue_comment_flag": False,
                "folder_metadata_match_note": "No metadata match found.",
            }

        trial = int(w["trial"])
        is_practice = bool(w["is_practice"])
        condition_map = expand_trial_order(meta.get("trial_order", ""))
        condition = "practice" if is_practice else condition_map.get(trial, "unknown")

        exclude_reasons: list[str] = []
        if is_practice:
            exclude_reasons.append("practice")
        if not bool(w["usable_for_dyadic_ecg"]):
            exclude_reasons.append("not_usable_for_dyadic_ecg")
        if EXCLUDE_PILOTS_FROM_MAIN and bool(meta.get("is_pilot", False)):
            exclude_reasons.append("pilot")
        if condition == "unknown":
            exclude_reasons.append("condition_unknown")
        if EXCLUDE_COMMENT_FLAGS_FROM_MAIN and (
            bool(meta.get("ball_drop_flag", False))
            or bool(meta.get("stopped_flag", False))
            or bool(meta.get("wrong_way_flag", False))
        ):
            exclude_reasons.append("comment_flag_excluded")

        rows.append(
            {
                **meta,
                "candidate_window": int(w["candidate_window"]),
                "trial": trial,
                "is_practice": is_practice,
                "condition": condition,
                "leader_role": "unknown",
                "accel_group_id": w.get("accel_group_id", np.nan),
                "timestamp_id": w.get("timestamp_id", np.nan),
                "table_timestamps_row_index": w.get("table_timestamps_row_index", np.nan),
                "accel_start_raw": w.get("accel_start_raw", ""),
                "accel_end_raw": w.get("accel_end_raw", ""),
                "accel_start_local": w.get("accel_start_local", ""),
                "accel_end_local": w.get("accel_end_local", ""),
                "accel_start_unix": w.get("accel_start_unix", np.nan),
                "accel_end_unix": w.get("accel_end_unix", np.nan),
                "window_duration_sec": w.get("window_duration_sec", np.nan),
                "usable_for_dyadic_ecg": bool(w.get("usable_for_dyadic_ecg", False)),
                "covered_sensor_count": w.get("covered_sensor_count", np.nan),
                "n_windows_in_accel_group": w.get("n_windows_in_accel_group", np.nan),
                "exclude_from_main_analysis": len(exclude_reasons) > 0,
                "exclude_reason": "; ".join(exclude_reasons),
            }
        )

    analysis = pd.DataFrame(rows)

    # Merge behavioural and rating variables from aggregated_data.csv.
    aggregated_clean = clean_column_names(aggregated.copy())
    aggregated_clean = aggregated_clean.drop(
        columns=[c for c in aggregated_clean.columns if str(c).startswith("Unnamed")],
        errors="ignore",
    )

    if "pair1" in analysis.columns and "pair1" in aggregated_clean.columns:
        analysis = analysis.merge(
            aggregated_clean,
            how="left",
            on=["pair1", "trial"],
            suffixes=("", "_aggregated"),
        )

    if "table_motion_mean" in analysis.columns:
        analysis["aggregated_data_missing"] = analysis["table_motion_mean"].isna()
    else:
        analysis["aggregated_data_missing"] = True

    if "trial_duration_mean" in analysis.columns:
        analysis["duration_difference_sec"] = analysis["window_duration_sec"] - analysis["trial_duration_mean"]

    # Readable column order.
    preferred = [
        "recording_folder",
        "dyad_id",
        "pair1",
        "participant_A",
        "participant_B",
        "sensor_A",
        "sensor_B",
        "candidate_window",
        "trial",
        "is_practice",
        "condition",
        "leader_role",
        "accel_start_unix",
        "accel_end_unix",
        "accel_start_local",
        "accel_end_local",
        "window_duration_sec",
        "usable_for_dyadic_ecg",
        "covered_sensor_count",
        "is_pilot",
        "exclude_from_main_analysis",
        "exclude_reason",
        "session_comments",
        "ball_drop_flag",
        "stopped_flag",
        "talking_laughing_flag",
        "wrong_way_flag",
        "sensor_issue_comment_flag",
        "table_motion_mean",
        "trial_duration_mean",
        "Coordinated_mean",
        "Joint_control_mean",
        "Incontrol_leading_mean",
        "RA",
        "folder_metadata_match_note",
        "sensor_role_resolution_note",
        "aggregated_data_missing",
        "duration_difference_sec",
    ]
    ordered = [c for c in preferred if c in analysis.columns]
    remaining = [c for c in analysis.columns if c not in ordered]
    return analysis[ordered + remaining]


# =============================================================================
# Workbook writing and summary
# =============================================================================

def make_qc_summary(
    inventory: pd.DataFrame,
    ecg_qc: pd.DataFrame,
    session_summary: pd.DataFrame,
    accel_windows: pd.DataFrame,
    accel_groups: pd.DataFrame,
    accel_group_folder_matches: pd.DataFrame,
    candidate_windows: pd.DataFrame,
    metadata_dyads: pd.DataFrame,
    folder_metadata: pd.DataFrame,
    analysis_units: pd.DataFrame,
    unmatched_complete_sessions_without_windows: pd.DataFrame,
    unmatched_metadata_folders: pd.DataFrame,
    unmatched_accel_groups: pd.DataFrame,
) -> pd.DataFrame:
    """Create human-readable QC summary table."""
    rows: list[dict[str, Any]] = []

    def add(item: str, value: Any, note: str = "") -> None:
        rows.append({"item": item, "value": value, "note": note})

    add("movesense_sensor_rows", len(inventory), "One row per recording folder and sensor folder.")
    add("session_folders_found", len(session_summary), "Includes complete and empty session folders.")

    if "session_status" in session_summary.columns:
        for status, count in session_summary["session_status"].value_counts(dropna=False).items():
            add(f"session_status_{status}", int(count), "Count of sessions by QC status.")

    add("complete_movesense_sessions", int(session_summary["session_complete_for_movesense"].sum()), "Complete raw Movesense sessions.")
    add("empty_session_folders", int((session_summary["session_status"] == "empty_folder").sum()), "Empty Movesense session folders.")
    add("ecg_qc_rows", len(ecg_qc), "One row per recording folder and sensor.")
    add("inferred_accel_windows", len(accel_windows), "Rows where datatype == accel in table_timestamps.csv.")
    if "accel_time_repair_applied" in accel_windows.columns:
        add(
            "accel_time_repaired_rows",
            int(accel_windows["accel_time_repair_applied"].astype(bool).sum()),
            "Accel rows with a documented manual datetime repair before grouping/matching.",
        )
    add("inferred_accel_groups", len(accel_groups), "Session-level table accelerometer groups inferred from accel rows.")
    add("accel_groups_matched_to_folders", len(accel_group_folder_matches), "Accel groups matched to Movesense folders.")
    add("matched_candidate_windows_complete_sessions", len(candidate_windows), "Candidate windows for complete Movesense sessions.")
    add("analysis_units_rows", len(analysis_units), "One row per dyad-window in the master analysis-unit table.")

    if not candidate_windows.empty:
        add(
            "candidate_windows_usable_for_dyadic_ecg",
            int(candidate_windows["usable_for_dyadic_ecg"].sum()),
            "Windows where both participant ECG files cover the full accel window.",
        )
        add(
            "candidate_windows_not_usable_for_dyadic_ecg",
            int((~candidate_windows["usable_for_dyadic_ecg"].astype(bool)).sum()),
            "Windows lacking full dyadic ECG coverage.",
        )

    if not analysis_units.empty:
        add(
            "main_analysis_eligible_rows",
            int((~analysis_units["exclude_from_main_analysis"].astype(bool)).sum()),
            "Rows not excluded for practice, pilot, ECG coverage, or condition issues.",
        )
        add(
            "main_analysis_excluded_rows",
            int(analysis_units["exclude_from_main_analysis"].astype(bool).sum()),
            "Rows excluded from main analysis under current settings.",
        )
        add(
            "condition_unknown_rows",
            int((analysis_units["condition"] == "unknown").sum()),
            "Rows with missing FF/FE condition label.",
        )
        add(
            "pilot_rows",
            int(analysis_units["is_pilot"].astype(bool).sum()) if "is_pilot" in analysis_units.columns else 0,
            "Analysis-unit rows marked as pilot.",
        )
        add(
            "aggregated_data_missing_rows",
            int(analysis_units["aggregated_data_missing"].astype(bool).sum()) if "aggregated_data_missing" in analysis_units.columns else "",
            "Rows without matching aggregated behavioural/rating data.",
        )

    add("metadata_dyads_rows", len(metadata_dyads), "One row per dyad parsed from Table Task 2026 metadata.")
    add("folders_with_metadata_match", len(folder_metadata), "Movesense folders matched to metadata.")
    add("folders_without_metadata_match", len(unmatched_metadata_folders), "Movesense folders not matched to metadata.")
    add("complete_sessions_without_accel_windows", len(unmatched_complete_sessions_without_windows), "Complete Movesense sessions without matched accel windows.")
    add("accel_groups_without_folder_match", len(unmatched_accel_groups), "Accel groups without a nearby Movesense folder.")

    return pd.DataFrame(rows)


def safe_sheet_name(name: str) -> str:
    """Excel sheet names cannot exceed 31 characters."""
    return name[:31]


def coerce_identifier_columns_to_text(df: pd.DataFrame) -> pd.DataFrame:
    """Keep long IDs and suffixes as text, especially for Excel display.

    Sensor IDs are 12-digit identifiers. If Excel interprets them as numbers,
    it displays them as scientific notation. This function keeps sensor-related
    columns as strings while leaving numeric analysis columns unchanged.
    """
    out = df.copy()
    for col in out.columns:
        if col in TEXT_DISPLAY_COLUMNS or col.endswith("_sensor_ids") or "sensor_" in col:
            out[col] = out[col].map(lambda x: "" if pd.isna(x) else str(x))
    return out


def make_excel_safe(df: pd.DataFrame) -> pd.DataFrame:
    """Convert timezone-aware datetimes and Python datetime objects to strings.

    Excel cannot store timezone-aware datetime values. Most study times are
    already stored as strings in this script, but a few diagnostic columns may
    contain Python datetime objects. This function makes workbook writing safe.
    """
    out = coerce_identifier_columns_to_text(df)

    for col in out.columns:
        # Convert pandas timezone-aware datetime columns to readable strings.
        if isinstance(out[col].dtype, pd.DatetimeTZDtype):
            out[col] = out[col].astype(str)
            continue

        # Convert object columns containing datetime/date/time objects to strings.
        if out[col].dtype == "object":
            out[col] = out[col].map(
                lambda x: str(x) if isinstance(x, (datetime, date, time, pd.Timestamp)) else x
            )

    return out


def write_master_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """Write all QC tables to one Excel workbook."""
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df_to_write = make_excel_safe(df)
            df_to_write.to_excel(writer, sheet_name=safe_sheet_name(name), index=False)

        workbook = writer.book
        for ws in workbook.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            # Force sensor-related columns to Text format so Excel does not
            # display long IDs in scientific notation.
            header_to_col = {str(cell.value): cell.column for cell in ws[1] if cell.value is not None}
            for header, col_idx in header_to_col.items():
                if header in TEXT_DISPLAY_COLUMNS or header.endswith("_sensor_ids") or "sensor_" in header:
                    col_letter = get_column_letter(col_idx)
                    for cell in ws[col_letter]:
                        cell.number_format = "@"

            # Auto-width based on first 200 cells in each column.
            for column_cells in ws.columns:
                max_len = 0
                letter = column_cells[0].column_letter
                for cell in column_cells[:200]:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)


# =============================================================================
# Main driver
# =============================================================================

def main() -> None:
    """Run full metadata/QC preparation workflow."""
    args = parse_args()
    root = Path(args.root).expanduser().resolve()

    movesense_dir = find_required_path(root, ["Table MoveSense Sensor"], "Table MoveSense Sensor folder")
    metadata_path = find_metadata_path(root)
    table_timestamps_path = find_required_path(root, ["table_timestamps.csv"], "table_timestamps.csv")
    aggregated_path = find_required_path(root, ["aggregated_data.csv"], "aggregated_data.csv")

    out_dir = root / "qc_outputs"
    out_dir.mkdir(exist_ok=True)

    print("Preparing TableTask metadata and QC")
    print(f"Root: {root}")
    print(f"Movesense folder: {movesense_dir}")
    print(f"Metadata sheet: {metadata_path.name}")
    print(f"table_timestamps.csv: {table_timestamps_path.name}")
    print(f"aggregated_data.csv: {aggregated_path.name}")

    print("\n[1/7] Scanning Movesense folders and ECG coverage...")
    inventory, ecg_qc, session_summary = build_movesense_inventory(movesense_dir)

    print("[2/7] Inferring full accel-window datetimes from table_timestamps.csv...")
    inferred_accel_windows = infer_accel_windows(table_timestamps_path)
    accel_groups = summarize_accel_groups(inferred_accel_windows)

    print("[3/7] Matching accel groups to Movesense folders...")
    accel_group_folder_matches, unmatched_accel_groups = match_accel_groups_to_folders(
        accel_groups, session_summary
    )

    print("[4/7] Creating candidate-window ECG coverage QC...")
    candidate_window_coverage, unmatched_complete_sessions_without_windows = build_candidate_window_coverage(
        inferred_accel_windows,
        accel_group_folder_matches,
        session_summary,
        ecg_qc,
    )

    print("[5/7] Reading metadata and aggregated behavioural/rating data...")
    metadata_raw = read_metadata_sheet(metadata_path)
    aggregated = pd.read_csv(aggregated_path)
    aggregated = clean_column_names(aggregated)
    metadata_dyads, metadata_raw_blocks = build_metadata_dyads(metadata_raw, aggregated)

    print("[6/7] Matching metadata to Movesense folders and creating analysis units...")
    folder_metadata_matches, unmatched_metadata_folders, metadata_match_candidates = match_metadata_to_folders(
        metadata_dyads,
        session_summary,
    )
    analysis_units = build_analysis_units(
        candidate_window_coverage,
        folder_metadata_matches,
        aggregated,
    )

    print("[7/7] Writing QC outputs...")
    qc_summary = make_qc_summary(
        inventory=inventory,
        ecg_qc=ecg_qc,
        session_summary=session_summary,
        accel_windows=inferred_accel_windows,
        accel_groups=accel_groups,
        accel_group_folder_matches=accel_group_folder_matches,
        candidate_windows=candidate_window_coverage,
        metadata_dyads=metadata_dyads,
        folder_metadata=folder_metadata_matches,
        analysis_units=analysis_units,
        unmatched_complete_sessions_without_windows=unmatched_complete_sessions_without_windows,
        unmatched_metadata_folders=unmatched_metadata_folders,
        unmatched_accel_groups=unmatched_accel_groups,
    )

    master_xlsx = out_dir / "TableTask_Master_QC.xlsx"
    analysis_csv = out_dir / "veronica_analysis_units.csv"
    analysis_xlsx = out_dir / "veronica_analysis_units.xlsx"
    summary_txt = out_dir / "metadata_qc_summary.txt"

    sheets = {
        "qc_summary": qc_summary,
        "session_summary": session_summary,
        "movesense_inventory": inventory,
        "ecg_coverage_qc": ecg_qc,
        "inferred_accel_windows": inferred_accel_windows,
        "accel_groups": accel_groups,
        "accel_group_folder_matches": accel_group_folder_matches,
        "candidate_window_coverage": candidate_window_coverage,
        "metadata_raw": metadata_raw,
        "metadata_raw_blocks": metadata_raw_blocks,
        "metadata_dyads": metadata_dyads,
        "metadata_match_candidates": metadata_match_candidates,
        "folder_metadata_matches": folder_metadata_matches,
        "analysis_units": analysis_units,
        "unmatched_complete_sessions": unmatched_complete_sessions_without_windows,
        "unmatched_metadata_folders": unmatched_metadata_folders,
        "unmatched_accel_groups": unmatched_accel_groups,
    }

    write_master_workbook(master_xlsx, sheets)

    # Save CSV for scripts and XLSX for human review. The XLSX version keeps
    # long sensor IDs displayed as text; CSV has no native cell formatting.
    analysis_units_for_output = coerce_identifier_columns_to_text(analysis_units)
    analysis_units_for_output.to_csv(analysis_csv, index=False)
    write_master_workbook(analysis_xlsx, {"analysis_units": analysis_units_for_output})

    with open(summary_txt, "w", encoding="utf-8") as f:
        f.write("TableTask metadata/QC summary\n")
        f.write("============================\n\n")
        for _, row in qc_summary.iterrows():
            f.write(f"{row['item']}: {row['value']}\n")
            if str(row.get("note", "")).strip():
                f.write(f"  Note: {row['note']}\n")

    print("\nDone.")
    print(f"Master QC workbook: {master_xlsx}")
    print(f"Analysis units CSV: {analysis_csv}")
    print(f"Analysis units Excel: {analysis_xlsx}")
    print(f"Text summary: {summary_txt}")

    print("\nQuick QC summary:")
    print(qc_summary.to_string(index=False))

    print("\nImportant checks before ECG preprocessing:")
    print("1. Open qc_outputs/TableTask_Master_QC.xlsx")
    print("2. Check qc_summary, session_summary, folder_metadata_matches")
    print("3. Check candidate_window_coverage and analysis_units")
    print("4. Do not proceed until condition_unknown_rows is 0 or explained")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print("\nERROR: metadata/QC preparation failed.", file=sys.stderr)
        print(str(exc), file=sys.stderr)
        raise
